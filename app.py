import streamlit as st
import anthropic
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import json
import base64
import re
from datetime import datetime
import copy

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="EMP Quotation Estimator",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1a237e, #283593);
        color: white;
        padding: 1.5rem 2rem;
        border-radius: 10px;
        margin-bottom: 1.5rem;
    }
    .main-header h1 { color: white; margin: 0; font-size: 1.8rem; }
    .main-header p  { color: #90caf9; margin: 0.3rem 0 0 0; font-size: 0.9rem; }
    .step-card {
        background: white;
        border: 1px solid #e0e0e0;
        border-left: 4px solid #1565c0;
        border-radius: 8px;
        padding: 1rem 1.2rem;
        margin-bottom: 1rem;
    }
    .step-card h3 { color: #1565c0; margin: 0 0 0.5rem 0; font-size: 1rem; }
    .db-card {
        background: #f8f9fa;
        border: 1px solid #dee2e6;
        border-radius: 6px;
        padding: 0.8rem;
        margin: 0.4rem 0;
    }
    .success-box {
        background: #e8f5e9;
        border: 1px solid #4caf50;
        border-radius: 8px;
        padding: 1rem;
        margin: 0.5rem 0;
    }
    .metric-box {
        background: #e3f2fd;
        border-radius: 8px;
        padding: 0.8rem 1rem;
        text-align: center;
        border: 1px solid #90caf9;
    }
    .metric-box .val { font-size: 1.4rem; font-weight: bold; color: #1565c0; }
    .metric-box .lbl { font-size: 0.8rem; color: #666; }
    .stButton>button { border-radius: 6px; }
    div[data-testid="stExpander"] { border-radius: 8px; }
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
  <h1>⚡ EMP Electrical Quotation Estimator</h1>
  <p>Electro Metal Pressings (Pvt) Ltd &nbsp;|&nbsp; Automated panel board costing from SLD drawings</p>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
for key, default in {
    "pdf_pages": [],          # list of {name, base64}
    "costing_data": {},       # parsed from costing model xlsx
    "db_items": [],           # extracted DB items [{db_name, components:[]}]
    "project_info": {},       # customer / project meta
    "output_excel": None,     # bytes of generated xlsx
    "extraction_done": False,
    "generation_done": False,
}.items():
    if key not in st.session_state:
        st.session_state[key] = default

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def pdf_to_base64_pages(pdf_bytes: bytes) -> list[dict]:
    """Return list of {page_num, base64_data} for each PDF page."""
    import pypdf, io
    try:
        reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
        pages = []
        for i, page in enumerate(reader.pages):
            writer = pypdf.PdfWriter()
            writer.add_page(page)
            buf = io.BytesIO()
            writer.write(buf)
            buf.seek(0)
            pages.append({
                "page_num": i + 1,
                "base64_data": base64.standard_b64encode(buf.read()).decode(),
            })
        return pages
    except Exception as e:
        st.error(f"PDF read error: {e}")
        return []


def load_costing_model(xlsx_bytes: bytes) -> dict:
    """Extract component price list from the EMP Costing Model xlsx."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    price_db = {}
    
    # Read 'Costing' sheet – rows have: S/N | Comments | TotalQTY | TotalPrice | ... | ComponentType | Description | ... | UnitPrice
    if "Costing" in wb.sheetnames:
        ws = wb["Costing"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[3:]:  # skip header rows
            if not row or len(row) < 8:
                continue
            comp_type = str(row[5]).strip() if row[5] else ""
            description = str(row[6]).strip() if row[6] else ""
            # unit price is in column 8 (index 7)
            try:
                price = float(row[7]) if row[7] and str(row[7]) not in ("0", "None", "") else 0
            except (ValueError, TypeError):
                price = 0
            if description and description != "None" and price > 0:
                key = description.lower().strip()
                price_db[key] = {"description": description, "type": comp_type, "price": price}
    
    # Also read 'All Cost Prices' if present
    if "All Cost Prices" in wb.sheetnames:
        ws = wb["All Cost Prices"]
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 3:
                continue
            desc = str(row[0]).strip() if row[0] else ""
            try:
                price = float(row[1]) if row[1] else 0
            except (ValueError, TypeError):
                price = 0
            if desc and desc != "None" and price > 0:
                key = desc.lower().strip()
                if key not in price_db:
                    price_db[key] = {"description": desc, "type": "General", "price": price}
    
    wb.close()
    return price_db


def parse_costing_summary(xlsx_bytes: bytes) -> dict:
    """Extract the Summary sheet parameters (margin, exchange rate, etc.)."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    params = {
        "margin": 0.40,
        "contingency": 0.05,
        "usd_buying": 315,
        "usd_selling": 335,
        "busbar_lme": 13800,
    }
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[:10]:
            if not row:
                continue
            for i, cell in enumerate(row):
                if cell == "Margin" and i + 1 < len(row):
                    try: params["margin"] = float(row[i+1])
                    except: pass
                if "Contingency" in str(cell) and i + 1 < len(row):
                    try: params["contingency"] = float(row[i+1])
                    except: pass
    wb.close()
    return params


def extract_db_from_pdfs(pdf_pages_all: list[dict]) -> tuple[list, dict]:
    """Use Claude vision to extract all distribution boards and their components from SLD PDFs."""
    client = anthropic.Anthropic()
    
    # Build the vision content
    content = []
    content.append({
        "type": "text",
        "text": """You are an expert electrical engineer reading Single Line Diagrams (SLD) for panel board estimation.

Analyze these SLD PDF pages and extract ALL distribution boards / panels shown.

For EACH board/panel extract:
1. The DB/Panel Name (e.g., MDB, TCU-GF-1, SDB-1F, CU-GF-2, etc.)
2. The incoming supply details (cable size, MCB/MCCB rating)
3. ALL protective devices with exact ratings: MCB (A, poles, kA), MCCB, RCCB (A, poles, mA), Isolators, ELR, SPD, etc.
4. Count of each device type

Return a JSON object with this EXACT structure:
{
  "panels": [
    {
      "name": "MDB",
      "incoming_cable": "120mm² 4x1C CU/XLPE/PVC",
      "incoming_breaker": "200A MCCB 4P 36kA",
      "components": [
        {"description": "10A 1P MCB 10kA Acti9", "qty": 12, "type": "MCB_Acti9_10kA"},
        {"description": "16A 1P MCB 10kA Acti9", "qty": 8, "type": "MCB_Acti9_10kA"},
        {"description": "32A 2P MCB 10kA Acti9", "qty": 3, "type": "MCB_Acti9_10kA"},
        {"description": "63A 4P MCB 10kA Acti9", "qty": 1, "type": "MCB_Acti9_10kA"},
        {"description": "RCCB 2P 40A 30mA", "qty": 2, "type": "RCCB_Acti9_10kA"},
        {"description": "RCCB 4P 40A 30mA", "qty": 1, "type": "RCCB_Acti9_10kA"},
        {"description": "200A 4P MCCB 36kA", "qty": 1, "type": "Compact_MCCB"},
        {"description": "SPD Class II", "qty": 1, "type": "Accessories"},
        {"description": "Enclosure / Panel Body", "qty": 1, "type": "Enclosures"}
      ]
    }
  ],
  "project_info": {
    "customer": "",
    "project": "",
    "description": "Supply of Panel Boards"
  }
}

IMPORTANT RULES:
- Count ALL MCBs carefully - they are the most numerous items
- Separate MCBs by: ampere rating AND pole count (1P, 2P, 3P, 4P)
- RCCBs: separate by poles (2P/4P), ampere rating, and mA sensitivity (30mA/100mA/300mA)
- MCCBs: note ampere rating, poles, and breaking capacity
- Include SPD, ELR, ATS controllers, isolators, contactors if shown
- Always include 1x Enclosure per panel
- Return ONLY valid JSON, no markdown code fences"""
    })
    
    # Add each PDF page as a document
    for page_info in pdf_pages_all[:15]:  # limit to 15 pages max
        content.append({
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": page_info["base64_data"],
            }
        })
    
    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=8000,
        messages=[{"role": "user", "content": content}]
    )
    
    raw = response.content[0].text
    # Strip any accidental markdown fences
    raw = re.sub(r"```json\s*", "", raw)
    raw = re.sub(r"```\s*", "", raw)
    raw = raw.strip()
    
    data = json.loads(raw)
    panels = data.get("panels", [])
    project_info = data.get("project_info", {})
    return panels, project_info


def match_price(description: str, price_db: dict) -> float:
    """Fuzzy-match a component description to the price database."""
    desc_lower = description.lower().strip()
    
    # 1. Exact match
    if desc_lower in price_db:
        return price_db[desc_lower]["price"]
    
    # 2. Parse MCB pattern: e.g. "10A 1P MCB 10kA"
    mcb_match = re.search(r"(\d+)a\s+(\d+)p\s+mcb\s+(\d+)ka", desc_lower)
    if mcb_match:
        amps, poles, ka = mcb_match.groups()
        # Search for matching entry
        for key, val in price_db.items():
            if f"{amps}a {poles}p mcb {ka}ka" in key or f"{amps}a {poles}p mcb" in key:
                return val["price"]
    
    # 3. Parse RCCB pattern
    rccb_match = re.search(r"rccb\s+(\d+)p\s+(\d+)a\s+(\d+)ma", desc_lower)
    if rccb_match:
        poles, amps, ma = rccb_match.groups()
        for key, val in price_db.items():
            if f"rccb {poles}p {amps}a {ma}ma" in key:
                return val["price"]
    
    # 4. Keyword scoring
    desc_words = set(desc_lower.split())
    best_score = 0
    best_price = 0
    for key, val in price_db.items():
        key_words = set(key.split())
        common = desc_words & key_words
        if len(common) >= 2:
            score = len(common) / max(len(desc_words), len(key_words))
            if score > best_score:
                best_score = score
                best_price = val["price"]
    
    return best_price if best_score > 0.4 else 0


def estimate_enclosure_cost(panel_name: str, component_count: int) -> float:
    """Rough enclosure cost based on component count."""
    # Very rough estimate: larger panels need bigger enclosures
    if component_count <= 10:
        return 45000   # Small CU
    elif component_count <= 20:
        return 75000   # Medium TCU
    elif component_count <= 35:
        return 120000  # Large TCU/SDB
    else:
        return 200000  # MDB / large panel


def calculate_db_cost(panel: dict, price_db: dict, params: dict) -> dict:
    """Calculate total cost for one panel."""
    components = panel.get("components", [])
    line_items = []
    material_cost = 0
    
    for comp in components:
        desc = comp.get("description", "")
        qty = int(comp.get("qty", 1))
        
        if "enclosure" in desc.lower() or "panel body" in desc.lower():
            unit_price = estimate_enclosure_cost(panel.get("name", ""), len(components))
        else:
            unit_price = match_price(desc, price_db)
        
        total = unit_price * qty
        material_cost += total
        line_items.append({
            "description": desc,
            "qty": qty,
            "unit_price": unit_price,
            "total": total,
            "matched": unit_price > 0,
        })
    
    # Add wiring / labour (typically 15-20% of material)
    wiring_cost = material_cost * 0.18
    
    # Contingency
    contingency = (material_cost + wiring_cost) * params["contingency"]
    
    # Total cost
    total_cost = material_cost + wiring_cost + contingency
    
    # Selling price with margin
    margin = params["margin"]
    selling_price = total_cost / (1 - margin) if margin < 1 else total_cost * 2
    
    # Round to nearest 10
    selling_price = round(selling_price / 10) * 10
    
    return {
        "name": panel.get("name", ""),
        "line_items": line_items,
        "material_cost": material_cost,
        "wiring_cost": wiring_cost,
        "contingency": contingency,
        "total_cost": total_cost,
        "selling_price": selling_price,
        "margin": margin,
    }


def generate_quotation_excel(
    project_info: dict,
    db_costs: list[dict],
    params: dict,
    template_bytes: bytes,
) -> bytes:
    """Generate the output quotation Excel, modelled on the EMP template."""
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes), keep_vba=False)
    
    # ── SUMMARY sheet ─────────────────────────────────────────────────────────
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        # Write DB list starting at row 5 (based on template structure)
        start_row = 5
        for i, db in enumerate(db_costs):
            row = start_row + i
            ws.cell(row=row, column=2, value=i + 1)           # S/N
            ws.cell(row=row, column=3, value=db["name"])       # DB Name
            ws.cell(row=row, column=4, value=1)                # QTY
            ws.cell(row=row, column=5, value=db["selling_price"])  # Unit Price
            ws.cell(row=row, column=6, value=db["selling_price"])  # Total
    
    # ── OFFER sheet ───────────────────────────────────────────────────────────
    if "Offer" in wb.sheetnames:
        ws = wb["Offer"]
        # Project info
        ref_no = project_info.get("ref_no", f"EMP/US/Q/{datetime.now().year}/")
        # Find and update key cells
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "EMP/US/Q/" in str(cell.value):
                    cell.value = ref_no
                    break
        
        # Write DB items in the offer table
        # Items start around row 9 in the template
        item_row = None
        for row in ws.iter_rows(min_row=1, max_row=20):
            for cell in row:
                if cell.value and "S/N" in str(cell.value):
                    item_row = cell.row + 1
                    break
            if item_row:
                break
        
        if not item_row:
            item_row = 9  # fallback
        
        total_amount = sum(db["selling_price"] for db in db_costs)
        
        for i, db in enumerate(db_costs):
            r = item_row + i
            # Find columns for S/N, Description, Qty, Unit, Rate, Total
            ws.cell(row=r, column=2, value=i + 1)
            ws.cell(row=r, column=3, value=db["name"])
            ws.cell(row=r, column=7, value=1)
            ws.cell(row=r, column=8, value="Nos")
            ws.cell(row=r, column=9, value=db["selling_price"])
            ws.cell(row=r, column=10, value=db["selling_price"])
    
    # ── BOM sheet — detailed breakdown ────────────────────────────────────────
    if "BOM" in wb.sheetnames:
        ws_bom = wb["BOM"]
    else:
        ws_bom = wb.create_sheet("BOM")
    
    # Clear and write BOM
    ws_bom.delete_rows(1, ws_bom.max_row)
    
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1565C0")
    hdr_align = Alignment(horizontal="center", vertical="center")
    
    headers = ["Panel", "S/N", "Description", "Qty", "Unit Price (LKR)", "Total (LKR)", "Matched"]
    for col, h in enumerate(headers, 1):
        c = ws_bom.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
    
    row_idx = 2
    panel_fills = ["F3F4F6", "FFFFFF"]
    for pi, db in enumerate(db_costs):
        fill = PatternFill("solid", fgColor=panel_fills[pi % 2])
        for li, item in enumerate(db["line_items"]):
            ws_bom.cell(row=row_idx, column=1, value=db["name"] if li == 0 else "").fill = fill
            ws_bom.cell(row=row_idx, column=2, value=li + 1)
            ws_bom.cell(row=row_idx, column=3, value=item["description"])
            ws_bom.cell(row=row_idx, column=4, value=item["qty"])
            ws_bom.cell(row=row_idx, column=5, value=item["unit_price"] if item["unit_price"] else "N/A")
            ws_bom.cell(row=row_idx, column=6, value=item["total"])
            ws_bom.cell(row=row_idx, column=7, value="✓" if item["matched"] else "⚠ Review")
            if not item["matched"]:
                ws_bom.cell(row=row_idx, column=7).font = Font(color="FF6600")
            row_idx += 1
        
        # Panel subtotal row
        sub_cell = ws_bom.cell(row=row_idx, column=3, value=f"PANEL TOTAL — {db['name']} (incl. margin)")
        sub_cell.font = Font(bold=True)
        ws_bom.cell(row=row_idx, column=6, value=db["selling_price"]).font = Font(bold=True, color="1565C0")
        row_idx += 2
    
    # Grand total
    ws_bom.cell(row=row_idx, column=3, value="GRAND TOTAL").font = Font(bold=True, size=12)
    ws_bom.cell(row=row_idx, column=6, value=sum(db["selling_price"] for db in db_costs)).font = Font(bold=True, size=12, color="1565C0")
    
    # Column widths
    ws_bom.column_dimensions["A"].width = 18
    ws_bom.column_dimensions["C"].width = 45
    ws_bom.column_dimensions["D"].width = 8
    ws_bom.column_dimensions["E"].width = 18
    ws_bom.column_dimensions["F"].width = 18
    ws_bom.column_dimensions["G"].width = 12
    
    # ── COST SUMMARY sheet ────────────────────────────────────────────────────
    if "Cost Summary" not in wb.sheetnames:
        ws_cs = wb.create_sheet("Cost Summary")
    else:
        ws_cs = wb["Cost Summary"]
    ws_cs.delete_rows(1, ws_cs.max_row)
    
    ws_cs["A1"] = "EMP Cost Summary"
    ws_cs["A1"].font = Font(bold=True, size=14)
    ws_cs["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_cs["A2"].font = Font(italic=True, color="666666")
    
    cs_headers = ["Panel", "Material Cost", "Wiring/Labour", "Contingency", "Total Cost", "Selling Price", "Gross Margin %"]
    for col, h in enumerate(cs_headers, 1):
        c = ws_cs.cell(row=4, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1565C0")
        c.alignment = Alignment(horizontal="center")
    
    for i, db in enumerate(db_costs):
        r = 5 + i
        ws_cs.cell(row=r, column=1, value=db["name"])
        ws_cs.cell(row=r, column=2, value=round(db["material_cost"]))
        ws_cs.cell(row=r, column=3, value=round(db["wiring_cost"]))
        ws_cs.cell(row=r, column=4, value=round(db["contingency"]))
        ws_cs.cell(row=r, column=5, value=round(db["total_cost"]))
        ws_cs.cell(row=r, column=6, value=db["selling_price"])
        gm = (db["selling_price"] - db["total_cost"]) / db["selling_price"] if db["selling_price"] else 0
        ws_cs.cell(row=r, column=7, value=round(gm * 100, 1))
    
    for col in range(1, 8):
        ws_cs.column_dimensions[get_column_letter(col)].width = 18
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## 📋 Project Details")
    
    proj_ref = st.text_input("Quotation Ref No.", placeholder="EMP/US/Q/26/XXXXX")
    proj_customer = st.text_input("Customer Name", placeholder="e.g. Synex International (Pvt) Ltd")
    proj_attention = st.text_input("Attention (Contact)", placeholder="e.g. Ms. Harshani Bendarage")
    proj_project = st.text_input("Project Name / Location", placeholder="e.g. NC")
    proj_desc = st.selectbox("Description", [
        "Supply of Panel Boards",
        "Supply of ATS Panel",
        "Supply of Capacitor Bank",
        "Supply of Synch Panel",
        "Supply of Pump Panel",
        "Supply of Panel Boards & ATS Panel",
    ])
    proj_sales = st.selectbox("Sales Representative", [
        "Kaushalya", "Kushan", "John", "Suneth", "Kavith", "Indika"
    ])
    proj_engineer = st.selectbox("Estimation Engineer", [
        "Uthpala", "Samith", "Nadeesha", "Aloka", "Rashmika"
    ])
    
    st.markdown("---")
    st.markdown("## ⚙️ Costing Parameters")
    margin = st.slider("Gross Margin %", 20, 60, 40) / 100
    contingency = st.slider("Contingency %", 0, 15, 5) / 100
    usd_buying = st.number_input("USD Buying Rate (LKR)", value=315)
    usd_selling = st.number_input("USD Selling Rate (LKR)", value=335)
    
    st.markdown("---")
    st.caption("⚡ EMP Quotation Estimator v1.0")

# Store sidebar params
sidebar_params = {
    "margin": margin,
    "contingency": contingency,
    "usd_buying": usd_buying,
    "usd_selling": usd_selling,
    "ref_no": proj_ref,
    "customer": proj_customer,
    "attention": proj_attention,
    "project": proj_project,
    "description": proj_desc,
}

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Upload Files
# ══════════════════════════════════════════════════════════════════════════════
st.markdown('<div class="step-card"><h3>📁 Step 1 — Upload Design PDFs & Costing Model</h3></div>', unsafe_allow_html=True)

col_a, col_b = st.columns([3, 2])

with col_a:
    sld_files = st.file_uploader(
        "Upload SLD Drawing PDFs (Single Line Diagrams)",
        type=["pdf"],
        accept_multiple_files=True,
        help="Upload all SLD PDF sets for the project"
    )

with col_b:
    costing_file = st.file_uploader(
        "Upload EMP Costing Model Excel (.xlsm / .xlsx)",
        type=["xlsm", "xlsx"],
        help="Upload the latest Quotation_EMP_Costing_Model xlsx"
    )

# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Extract from SLDs
# ══════════════════════════════════════════════════════════════════════════════
st.markdown('<div class="step-card"><h3>🔍 Step 2 — Extract Distribution Boards from SLD</h3></div>', unsafe_allow_html=True)

if sld_files and costing_file:
    if st.button("⚡ Extract Components from SLD Drawings", type="primary", use_container_width=True):
        with st.spinner("Reading PDF pages…"):
            all_pages = []
            for f in sld_files:
                pages = pdf_to_base64_pages(f.read())
                for p in pages:
                    p["filename"] = f.name
                all_pages.extend(pages)
            st.session_state["pdf_pages"] = all_pages
        
        with st.spinner("Loading costing model…"):
            price_db = load_costing_model(costing_file.read())
            costing_file.seek(0)
            template_bytes = costing_file.read()
            st.session_state["costing_data"] = price_db
            st.session_state["template_bytes"] = template_bytes
            params_from_file = parse_costing_summary(template_bytes)
            st.session_state["file_params"] = params_from_file
        
        with st.spinner(f"Analysing {len(all_pages)} SLD page(s) with Claude AI…"):
            try:
                panels, project_info = extract_db_from_pdfs(all_pages)
                st.session_state["db_items"] = panels
                st.session_state["project_info"] = project_info
                st.session_state["extraction_done"] = True
                st.success(f"✅ Extracted {len(panels)} distribution boards / panels!")
            except Exception as e:
                st.error(f"Extraction failed: {e}")

elif not sld_files:
    st.info("👆 Please upload SLD PDF files to continue")
elif not costing_file:
    st.info("👆 Please upload the EMP Costing Model Excel file to continue")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Review & Edit
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state["extraction_done"] and st.session_state["db_items"]:
    st.markdown('<div class="step-card"><h3>✏️ Step 3 — Review & Edit Extracted Components</h3></div>', unsafe_allow_html=True)
    
    panels = st.session_state["db_items"]
    price_db = st.session_state["costing_data"]
    
    # Summary metrics
    total_mcbs = sum(
        sum(c.get("qty", 0) for c in p.get("components", []) if "mcb" in c.get("description", "").lower())
        for p in panels
    )
    total_rccbs = sum(
        sum(c.get("qty", 0) for c in p.get("components", []) if "rccb" in c.get("description", "").lower())
        for p in panels
    )
    
    m1, m2, m3 = st.columns(3)
    with m1:
        st.markdown(f'<div class="metric-box"><div class="val">{len(panels)}</div><div class="lbl">Panels Detected</div></div>', unsafe_allow_html=True)
    with m2:
        st.markdown(f'<div class="metric-box"><div class="val">{total_mcbs}</div><div class="lbl">Total MCBs</div></div>', unsafe_allow_html=True)
    with m3:
        st.markdown(f'<div class="metric-box"><div class="val">{total_rccbs}</div><div class="lbl">Total RCCBs</div></div>', unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Editable panel list
    updated_panels = []
    for pi, panel in enumerate(panels):
        with st.expander(f"📦 {panel.get('name', f'Panel {pi+1}')} — {len(panel.get('components', []))} items", expanded=(pi == 0)):
            panel_name = st.text_input("Panel Name", value=panel.get("name", ""), key=f"pname_{pi}")
            
            components = panel.get("components", [])
            updated_comps = []
            
            # Display components in a table-like format
            if components:
                comp_df_data = []
                for ci, comp in enumerate(components):
                    desc = st.text_input(f"Item {ci+1} Description", value=comp.get("description", ""), key=f"desc_{pi}_{ci}", label_visibility="collapsed")
                    col1, col2, col3 = st.columns([4, 1, 2])
                    with col1:
                        st.caption(desc[:60] if len(desc) > 60 else desc)
                    with col2:
                        qty = st.number_input("Qty", min_value=0, value=int(comp.get("qty", 1)), key=f"qty_{pi}_{ci}", label_visibility="collapsed")
                    with col3:
                        matched_price = match_price(desc, price_db)
                        if matched_price > 0:
                            st.caption(f"✓ LKR {matched_price:,.0f}")
                        else:
                            st.caption("⚠️ No price match")
                    
                    if qty > 0:
                        updated_comps.append({"description": desc, "qty": qty, "type": comp.get("type", "")})
            
            # Add component button
            if st.button(f"➕ Add Component", key=f"add_{pi}"):
                updated_comps.append({"description": "New Component", "qty": 1, "type": "Accessories"})
            
            updated_panels.append({"name": panel_name, "components": updated_comps,
                                    "incoming_cable": panel.get("incoming_cable", ""),
                                    "incoming_breaker": panel.get("incoming_breaker", "")})
    
    st.session_state["db_items"] = updated_panels

    # ══════════════════════════════════════════════════════════════════════════
    # STEP 4 — Calculate & Generate
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="step-card"><h3>🧮 Step 4 — Calculate Costs & Generate Quotation</h3></div>', unsafe_allow_html=True)
    
    params = {**sidebar_params, "contingency": contingency}
    
    if st.button("🚀 Calculate & Generate Quotation Excel", type="primary", use_container_width=True):
        price_db = st.session_state["costing_data"]
        panels = st.session_state["db_items"]
        template_bytes = st.session_state.get("template_bytes", b"")
        
        with st.spinner("Calculating costs for each panel…"):
            db_costs = []
            for panel in panels:
                if panel.get("name"):
                    cost = calculate_db_cost(panel, price_db, params)
                    db_costs.append(cost)
        
        # Show cost summary table
        st.markdown("#### 💰 Cost Summary")
        summary_data = []
        for db in db_costs:
            gm = (db["selling_price"] - db["total_cost"]) / db["selling_price"] * 100 if db["selling_price"] else 0
            summary_data.append({
                "Panel": db["name"],
                "Material (LKR)": f"{db['material_cost']:,.0f}",
                "Labour (LKR)": f"{db['wiring_cost']:,.0f}",
                "Total Cost (LKR)": f"{db['total_cost']:,.0f}",
                "Selling Price (LKR)": f"{db['selling_price']:,.0f}",
                "GM %": f"{gm:.1f}%",
            })
        
        df = pd.DataFrame(summary_data)
        st.dataframe(df, use_container_width=True, hide_index=True)
        
        grand_total = sum(db["selling_price"] for db in db_costs)
        vat = grand_total * 0.15
        sscl = grand_total * 0.025
        
        st.markdown(f"""
        <div class="success-box">
          <b>Grand Total (excl. taxes):</b> LKR {grand_total:,.0f}<br>
          <b>SSCL (2.5%):</b> LKR {sscl:,.0f}<br>
          <b>VAT (15%):</b> LKR {vat:,.0f}<br>
          <b>Grand Total (incl. taxes):</b> LKR {grand_total + sscl + vat:,.0f}
        </div>
        """, unsafe_allow_html=True)
        
        # Generate Excel
        with st.spinner("Generating quotation Excel file…"):
            proj_info = {
                "ref_no": proj_ref or f"EMP/US/Q/{datetime.now().year}/",
                "customer": proj_customer or st.session_state["project_info"].get("customer", ""),
                "attention": proj_attention,
                "project": proj_project or st.session_state["project_info"].get("project", ""),
                "description": proj_desc,
            }
            
            output_bytes = generate_quotation_excel(proj_info, db_costs, params, template_bytes)
            st.session_state["output_excel"] = output_bytes
            st.session_state["generation_done"] = True

# ══════════════════════════════════════════════════════════════════════════════
# STEP 5 — Download
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state["generation_done"] and st.session_state["output_excel"]:
    st.markdown('<div class="step-card"><h3>📥 Step 5 — Download Quotation</h3></div>', unsafe_allow_html=True)
    
    fname = f"EMP_Quotation_{proj_customer.replace(' ', '_') if proj_customer else 'Output'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    st.download_button(
        label="⬇️ Download Quotation Excel",
        data=st.session_state["output_excel"],
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )
    
    st.info("The Excel file contains: **Offer** · **Cost Summary** · **BOM** sheets based on the EMP Costing Model template.")

# ══════════════════════════════════════════════════════════════════════════════
# Footer
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("---")
st.caption("⚡ EMP Electrical Quotation Estimator | Powered by Claude AI | Electro Metal Pressings (Pvt) Ltd")
